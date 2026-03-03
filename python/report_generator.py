#!/usr/bin/env python3
"""
BuildFileTracker Report Generator
Converts raw tracking data into various report formats
"""

import json
import csv
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import sys
import os
import platform
import subprocess
import shutil
import re

class ReportGenerator:
    """Generate reports from build file tracking data"""
    
    # Mapping of file extensions to linking types
    STATIC_LIBRARY_EXTENSIONS = {'.a', '.lib'}  # .lib can be static
    SHARED_LIBRARY_EXTENSIONS = {'.so', '.dll', '.dylib'}
    OBJECT_FILE_EXTENSIONS = {'.o', '.obj'}
    PYTHON_EXTENSION_EXTENSIONS = {'.pyd', '.so'}  # .pyd on Windows, .so on Linux/macOS
    SOURCE_EXTENSIONS = {'.c', '.cc', '.cpp', '.cxx', '.m', '.mm', '.rs', '.go', '.java', '.kt', '.swift', '.cs', '.py', '.js', '.ts'}
    HEADER_EXTENSIONS = {'.h', '.hh', '.hpp', '.hxx', '.inl'}
    ARCHIVE_EXTENSIONS = {'.a', '.lib', '.zip', '.tar', '.tgz', '.tar.gz', '.tar.bz2', '.7z', '.rar', '.jar', '.war', '.ear', '.whl', '.egg'}
    BINARY_EXTENSIONS = {'.exe', '.bin'}
    CONFIG_EXTENSIONS = {'.json', '.xml', '.yml', '.yaml', '.toml', '.ini', '.cfg', '.conf', '.properties'}
    RESOURCE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.svg', '.ico', '.bmp', '.ttf', '.otf', '.woff', '.woff2', '.mp3', '.wav', '.ogg', '.mp4', '.mkv', '.avi'}
    
    def __init__(self, input_json: str):
        """Initialize with input JSON file"""
        self.input_file = input_json
        self.data = self._load_data()
    
    def _load_data(self) -> Dict:
        """Load data from JSON file"""
        try:
            with open(self.input_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Error: Input file '{self.input_file}' not found")
            sys.exit(1)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON in input file: {e}")
            sys.exit(1)
    
    @staticmethod
    def detect_linking_type(filepath: str, file_type: str = None) -> str:
        """
        Detect the linking type for a file
        
        Returns:
            - "static" for static libraries (.a, .lib) and object files
            - "dynamic" for shared libraries (.so, .dll, .dylib)
            - "import" for Windows import libraries (.lib for .dll)
            - "python_extension" for Python compiled extensions (.pyd, .so with Python in path)
            - "framework" for macOS frameworks
            - "runtime_plugin" for plugins in standard plugin directories
            - "unknown" if cannot determine
        """
        # Get file extension
        ext = Path(filepath).suffix.lower()
        
        if not file_type:
            file_type = ext.lstrip('.')
        
        # Framework linking (macOS)
        if '.framework' in filepath.lower():
            return "framework"
        
        # Python extensions (in site-packages)
        if 'site-packages' in filepath or 'dist-packages' in filepath:
            if ext in {'.so', '.pyd'}:
                return "python_extension"
        
        # Runtime plugins (standard plugin directories)
        plugin_paths = ['/usr/lib/plugins', '/usr/local/lib/plugins', 
                       'plugins', '.so', 'plugin', '.dll']
        for plugin_path in plugin_paths:
            if plugin_path in filepath.lower():
                if ext in ReportGenerator.SHARED_LIBRARY_EXTENSIONS:
                    return "runtime_plugin"
        
        # Object files
        if ext in ReportGenerator.OBJECT_FILE_EXTENSIONS:
            return "object"
        
        # Shared libraries (dynamic)
        if ext in ReportGenerator.SHARED_LIBRARY_EXTENSIONS:
            return "dynamic"
        
        # Static libraries
        if ext in ReportGenerator.STATIC_LIBRARY_EXTENSIONS:
            # Windows distinguishes import .lib from static .lib
            # Import libs are typically with .dll files in Windows/System
            if ext == '.lib':
                if 'Windows' in filepath or 'System' in filepath or 'SysWOW64' in filepath:
                    return "import"
            return "static"
        
        # Bitcode (LTO)
        if ext == '.bc':
            return "lto_bitcode"
        
        return "unknown"
    
    def enrich_data_with_linking_types(self):
        """Add linking type information to each file in the accessed_files list"""
        if 'accessed_files' not in self.data:
            return
        
        for file_entry in self.data['accessed_files']:
            filepath = file_entry.get('filepath', '')
            file_type = file_entry.get('file_type', '')
            
            # Add linking type if not already present
            if 'linking_type' not in file_entry:
                file_entry['linking_type'] = self.detect_linking_type(filepath, file_type)

    def _find_first_tool(self, tool_names: List[str]) -> str:
        for tool in tool_names:
            tool_path = shutil.which(tool)
            if tool_path:
                return tool_path
        return ""

    def _extract_source_paths_from_text(self, text: str) -> List[str]:
        # Extract likely source file paths from tool output
        source_exts = ('.c', '.cc', '.cpp', '.cxx', '.h', '.hpp', '.hh', '.m', '.mm', '.s', '.S', '.rs', '.go')
        paths = set()

        # Common DWARF outputs include quoted file names
        for match in re.findall(r"'([^']+)'", text):
            if match.lower().endswith(source_exts):
                paths.add(match)

        # Also scan whitespace-delimited tokens for file-like paths
        for token in re.split(r"\s+", text):
            if token.lower().endswith(source_exts):
                paths.add(token.strip())

        return sorted(paths)

    def _get_sources_from_object(self, filepath: str) -> List[str]:
        # Best-effort DWARF extraction of source files from .o/.obj
        if not os.path.exists(filepath):
            return []

        system = platform.system()
        tool_cmd = []

        if system == 'Darwin':
            tool = self._find_first_tool(['llvm-dwarfdump', 'dwarfdump'])
            if tool:
                tool_cmd = [tool, '--debug-line', filepath]
        elif system == 'Linux':
            tool = self._find_first_tool(['llvm-objdump', 'objdump', 'readelf'])
            if tool.endswith('readelf'):
                tool_cmd = [tool, '--debug-dump=decodedline', filepath]
            elif tool:
                tool_cmd = [tool, '--dwarf=decodedline', filepath]
        elif system == 'Windows':
            tool = self._find_first_tool(['llvm-dwarfdump', 'llvm-objdump'])
            if tool and tool.endswith('llvm-dwarfdump'):
                tool_cmd = [tool, '--debug-line', filepath]
            elif tool:
                tool_cmd = [tool, '--dwarf=decodedline', filepath]

        if not tool_cmd:
            return []

        try:
            result = subprocess.run(
                tool_cmd,
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode != 0:
                return []
            return self._extract_source_paths_from_text(result.stdout)
        except Exception:
            return []

    def enrich_data_with_source_mapping(self):
        """Add source_path/source_paths to object file entries when possible."""
        if 'accessed_files' not in self.data:
            return

        for file_entry in self.data['accessed_files']:
            filepath = file_entry.get('filepath', '')
            file_type = file_entry.get('file_type', '')
            linking_type = file_entry.get('linking_type', '')

            is_object = linking_type == 'object' or file_type in ('o', 'obj')
            if not is_object:
                continue

            if 'source_path' in file_entry or 'source_paths' in file_entry:
                continue

            sources = self._get_sources_from_object(filepath)
            if sources:
                file_entry['source_paths'] = sources
                if len(sources) == 1:
                    file_entry['source_path'] = sources[0]

    def _classify_entry(self, file_entry: Dict) -> str:
        filepath = file_entry.get('filepath', '')
        ext = Path(filepath).suffix.lower()
        linking_type = file_entry.get('linking_type', '')

        if ext in self.SOURCE_EXTENSIONS:
            return 'sources'
        if ext in self.HEADER_EXTENSIONS:
            return 'headers'
        if linking_type == 'object' or ext in self.OBJECT_FILE_EXTENSIONS:
            return 'objects'
        if ext in self.SHARED_LIBRARY_EXTENSIONS or ext in self.STATIC_LIBRARY_EXTENSIONS:
            return 'libraries'
        if ext in self.BINARY_EXTENSIONS:
            return 'binaries'
        if ext in self.ARCHIVE_EXTENSIONS:
            return 'archives'
        if ext in self.CONFIG_EXTENSIONS:
            return 'configs'
        if ext in self.RESOURCE_EXTENSIONS:
            return 'resources'
        return 'others'

    def generate_component_report(self, output_file: str):
        """Generate a component-wise JSON report grouped by package name."""
        self.enrich_data_with_linking_types()
        self.enrich_data_with_source_mapping()

        components: Dict[str, Dict[str, List[str]]] = {}

        for file_entry in self.data.get('accessed_files', []):
            component = file_entry.get('package', 'unknown') or 'unknown'
            group = self._classify_entry(file_entry)
            path = file_entry.get('filepath', '')

            if component not in components:
                components[component] = {
                    'sources': [],
                    'headers': [],
                    'objects': [],
                    'libraries': [],
                    'binaries': [],
                    'archives': [],
                    'configs': [],
                    'resources': [],
                    'others': [],
                    'source_paths_from_objects': []
                }

            if path:
                components[component][group].append(path)

            if group == 'objects':
                sources = file_entry.get('source_paths', [])
                if sources:
                    components[component]['source_paths_from_objects'].extend(sources)

        # De-duplicate and sort
        for component, groups in components.items():
            for key, values in groups.items():
                groups[key] = sorted(set(values))

        report = {
            'report_type': 'build_file_tracker_components',
            'timestamp': str(self.data.get('timestamp', '')),
            'component_grouping': 'package',
            'components': components
        }

        with open(output_file, 'w') as f:
            json.dump(report, f, indent=2)
        print(f"Component report generated: {output_file}")
    
    def generate_json(self, output_file: str, pretty: bool = True):
        """Generate JSON report with linking type information"""
        # Enrich data with linking types
        self.enrich_data_with_linking_types()
        self.enrich_data_with_source_mapping()
        
        indent = 2 if pretty else None
        with open(output_file, 'w') as f:
            json.dump(self.data, f, indent=indent)
        print(f"JSON report generated: {output_file}")
    
    def generate_csv(self, output_file: str):
        """Generate CSV report with linking type information"""
        self.enrich_data_with_linking_types()
        self.enrich_data_with_source_mapping()
        
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Filepath', 'Package', 'File Type', 'Linking Type', 'Source Paths', 'Access Count'])
            
            for file_entry in self.data.get('accessed_files', []):
                writer.writerow([
                    file_entry.get('filepath', ''),
                    file_entry.get('package', ''),
                    file_entry.get('file_type', ''),
                    file_entry.get('linking_type', ''),
                    ';'.join(file_entry.get('source_paths', [])) or file_entry.get('source_path', ''),
                    file_entry.get('access_count', 0)
                ])
        print(f"CSV report generated: {output_file}")
    
    def generate_xml(self, output_file: str):
        """Generate XML report with linking type information"""
        self.enrich_data_with_linking_types()
        self.enrich_data_with_source_mapping()
        
        root = ET.Element('BuildFileTracker')
        
        metadata = ET.SubElement(root, 'metadata')
        ET.SubElement(metadata, 'report_type').text = self.data.get('report_type', 'build_file_tracker')
        ET.SubElement(metadata, 'timestamp').text = str(self.data.get('timestamp', ''))
        ET.SubElement(metadata, 'total_files').text = str(len(self.data.get('accessed_files', [])))
        
        files = ET.SubElement(root, 'accessed_files')
        
        for file_entry in self.data.get('accessed_files', []):
            file_elem = ET.SubElement(files, 'file')
            ET.SubElement(file_elem, 'filepath').text = file_entry.get('filepath', '')
            ET.SubElement(file_elem, 'package').text = file_entry.get('package', '')
            ET.SubElement(file_elem, 'file_type').text = file_entry.get('file_type', '')
            ET.SubElement(file_elem, 'linking_type').text = file_entry.get('linking_type', '')
            ET.SubElement(file_elem, 'source_paths').text = ';'.join(file_entry.get('source_paths', []))
            ET.SubElement(file_elem, 'access_count').text = str(file_entry.get('access_count', 0))
        
        # Pretty print XML
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
        with open(output_file, 'w') as f:
            f.write(xml_str)
        print(f"XML report generated: {output_file}")
    
    def generate_xlsx(self, output_file: str):
        """Generate Excel report with linking type information (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("Error: openpyxl library is required for Excel reports")
            print("Install it with: pip install openpyxl")
            sys.exit(1)
        
        self.enrich_data_with_linking_types()
        self.enrich_data_with_source_mapping()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "File Access Report"
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['Filepath', 'Package', 'File Type', 'Linking Type', 'Source Paths', 'Access Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row, file_entry in enumerate(self.data.get('accessed_files', []), 2):
            ws.cell(row=row, column=1, value=file_entry.get('filepath', ''))
            ws.cell(row=row, column=2, value=file_entry.get('package', ''))
            ws.cell(row=row, column=3, value=file_entry.get('file_type', ''))
            ws.cell(row=row, column=4, value=file_entry.get('linking_type', ''))
            ws.cell(row=row, column=5, value=';'.join(file_entry.get('source_paths', [])) or file_entry.get('source_path', ''))
            ws.cell(row=row, column=6, value=file_entry.get('access_count', 0))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15
        
        wb.save(output_file)
        print(f"Excel report generated: {output_file}")
    
    def generate_summary(self, output_file: str):
        """Generate a summary report with linking type breakdown"""
        self.enrich_data_with_linking_types()
        
        files = self.data.get('accessed_files', [])
        
        # Aggregate statistics
        packages = {}
        file_types = {}
        linking_types = {}
        total_accesses = 0
        
        for file_entry in files:
            pkg = file_entry.get('package', 'unknown')
            ftype = file_entry.get('file_type', 'unknown')
            ltype = file_entry.get('linking_type', 'unknown')
            count = file_entry.get('access_count', 0)
            
            packages[pkg] = packages.get(pkg, 0) + 1
            file_types[ftype] = file_types.get(ftype, 0) + 1
            linking_types[ltype] = linking_types.get(ltype, 0) + 1
            total_accesses += count
        
        with open(output_file, 'w') as f:
            f.write("=" * 70 + "\n")
            f.write("Build File Tracker - Summary Report\n")
            f.write("=" * 70 + "\n\n")
            
            f.write(f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Unique Files Accessed: {len(files)}\n")
            f.write(f"Total File Accesses: {total_accesses}\n\n")
            
            f.write("-" * 70 + "\n")
            f.write("Linking Type Breakdown:\n")
            f.write("-" * 70 + "\n")
            for ltype, count in sorted(linking_types.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / len(files)) * 100 if files else 0
                f.write(f"  {ltype:20s} : {count:5d} files ({percentage:6.2f}%)\n")
            
            f.write("\n" + "-" * 70 + "\n")
            f.write("Files by Package:\n")
            f.write("-" * 70 + "\n")
            for pkg, count in sorted(packages.items(), key=lambda x: x[1], reverse=True):
                f.write(f"  {pkg:40s} : {count:5d} files\n")
            
            f.write("\n" + "-" * 70 + "\n")
            f.write("Files by Type:\n")
            f.write("-" * 70 + "\n")
            for ftype, count in sorted(file_types.items(), key=lambda x: x[1], reverse=True):
                f.write(f"  {ftype:40s} : {count:5d} files\n")
            
            f.write("\n" + "=" * 70 + "\n")
        
        print(f"Summary report generated: {output_file}")
    
    def generate_package_report(self, package_name: str, output_file: str):
        """Generate a report for a specific package with linking type information"""
        self.enrich_data_with_linking_types()
        
        files = [f for f in self.data.get('accessed_files', []) 
                 if f.get('package') == package_name]
        
        if not files:
            print(f"No files found for package: {package_name}")
            return
        
        # Categorize by linking type
        linking_breakdown = {}
        for f in files:
            ltype = f.get('linking_type', 'unknown')
            if ltype not in linking_breakdown:
                linking_breakdown[ltype] = []
            linking_breakdown[ltype].append(f)
        
        with open(output_file, 'w') as f:
            f.write("=" * 70 + "\n")
            f.write(f"Build File Tracker - Package Report: {package_name}\n")
            f.write("=" * 70 + "\n\n")
            
            f.write(f"Total Files Accessed: {len(files)}\n")
            f.write(f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write("-" * 70 + "\n")
            f.write("Files by Linking Type:\n")
            f.write("-" * 70 + "\n")
            for ltype, flist in sorted(linking_breakdown.items(), key=lambda x: len(x[1]), reverse=True):
                f.write(f"\n{ltype.upper()} ({len(flist)} files):\n")
                for file_entry in sorted(flist, key=lambda x: x.get('access_count', 0), reverse=True):
                    f.write(f"  {file_entry.get('filepath', '')}\n")
                    f.write(f"    Access Count: {file_entry.get('access_count', 0)}\n")
        
        print(f"Package report generated: {output_file}")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate reports from build file tracking data'
    )
    parser.add_argument('input', help='Input JSON file from tracker')
    parser.add_argument('-o', '--output', help='Output file base name (without extension)')
    parser.add_argument('-f', '--format', 
                       choices=['json', 'csv', 'xml', 'xlsx', 'summary', 'all'],
                       default='all',
                       help='Output format (default: all)')
    parser.add_argument('-p', '--package', 
                       help='Generate report for specific package')
    parser.add_argument('--components', action='store_true',
                       help='Generate component-wise JSON report grouped by package')
    
    args = parser.parse_args()
    
    generator = ReportGenerator(args.input)
    
    if not args.output:
        args.output = Path(args.input).stem + '_report'
    
    if args.package:
        generator.generate_package_report(args.package, f"{args.output}_{args.package}.txt")
    elif args.format == 'all':
        generator.generate_json(f"{args.output}.json")
        generator.generate_csv(f"{args.output}.csv")
        generator.generate_xml(f"{args.output}.xml")
        generator.generate_summary(f"{args.output}_summary.txt")
        try:
            generator.generate_xlsx(f"{args.output}.xlsx")
        except SystemExit:
            print("Skipping XLSX generation (openpyxl not installed)")
        generator.generate_component_report(f"{args.output}_components.json")
    else:
        if args.format == 'json':
            generator.generate_json(f"{args.output}.json")
        elif args.format == 'csv':
            generator.generate_csv(f"{args.output}.csv")
        elif args.format == 'xml':
            generator.generate_xml(f"{args.output}.xml")
        elif args.format == 'xlsx':
            generator.generate_xlsx(f"{args.output}.xlsx")
        elif args.format == 'summary':
            generator.generate_summary(f"{args.output}_summary.txt")

    if args.components and args.format != 'all' and not args.package:
        generator.generate_component_report(f"{args.output}_components.json")


if __name__ == '__main__':
    main()

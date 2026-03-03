#!/usr/bin/env python3
"""
BuildFileTracker - Report Generation Module
Supports: JSON, CSV, XML, XLSX, SPDX, CycloneDX, Markdown, HTML

Copyright (c) 2026 BuildFileTracker Contributors
SPDX-License-Identifier: GPL-3.0-or-later
"""

import json
import csv
import uuid
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from buildfiletracker import BuildFileTracker, PackageInfo, ReportFormat

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


class ReportGenerator:
    """Generate reports in multiple formats"""
    
    def __init__(self, tracker: BuildFileTracker):
        self.tracker = tracker
    
    def generate(self, output_path: str, format: ReportFormat) -> None:
        """Generate report in specified format"""
        generators = {
            ReportFormat.JSON: self._generate_json,
            ReportFormat.CSV: self._generate_csv,
            ReportFormat.XML: self._generate_xml,
            ReportFormat.XLSX: self._generate_xlsx,
            ReportFormat.SPDX_JSON: self._generate_spdx_json,
            ReportFormat.SPDX_TAGVALUE: self._generate_spdx_tagvalue,
            ReportFormat.CYCLONEDX_JSON: self._generate_cyclonedx_json,
            ReportFormat.CYCLONEDX_XML: self._generate_cyclonedx_xml,
            ReportFormat.MARKDOWN: self._generate_markdown,
            ReportFormat.HTML: self._generate_html,
        }
        
        generator = generators.get(format)
        if not generator:
            raise ValueError(f"Unsupported format: {format}")
        
        generator(output_path)
        print(f"✅ Report generated: {output_path}")
    
    def _get_report_data(self) -> Dict:
        """Get base report data"""
        return {
            "buildfiletracker_version": "2.0.0",
            "generated_at": datetime.now().isoformat(),
            "backend": self.tracker.active_backend.value if self.tracker.active_backend else "none",
            "statistics": self.tracker.stats.to_dict(),
            "packages": [pkg.to_dict() for pkg in self.tracker.get_all_packages()]
        }
    
    # ========================================================================
    # JSON REPORT
    # ========================================================================
    
    def _generate_json(self, output_path: str) -> None:
        """Generate JSON report"""
        data = self._get_report_data()
        
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
    
    # ========================================================================
    # CSV REPORT
    # ========================================================================
    
    def _generate_csv(self, output_path: str) -> None:
        """Generate CSV report"""
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Package', 'Version', 'License', 'Total Files',
                'Used Files', 'Usage %', 'File Path'
            ])
            
            for pkg in self.tracker.get_all_packages():
                if pkg.used_files:
                    for filepath in sorted(pkg.used_files):
                        writer.writerow([
                            pkg.name, pkg.version, pkg.license,
                            pkg.total_files, pkg.used_files_count,
                            f"{pkg.usage_percentage:.2f}",
                            filepath
                        ])
                else:
                    writer.writerow([
                        pkg.name, pkg.version, pkg.license,
                        pkg.total_files, 0, "0.00", ""
                    ])
    
    # ========================================================================
    # XML REPORT
    # ========================================================================
    
    def _generate_xml(self, output_path: str) -> None:
        """Generate XML report"""
        data = self._get_report_data()
        
        root = ET.Element('buildfiletracker_report')
        root.set('version', data['buildfiletracker_version'])
        root.set('generated_at', data['generated_at'])
        root.set('backend', data['backend'])
        
        # Statistics
        stats_elem = ET.SubElement(root, 'statistics')
        for key, value in data['statistics'].items():
            elem = ET.SubElement(stats_elem, key)
            elem.text = str(value)
        
        # Packages
        packages_elem = ET.SubElement(root, 'packages')
        for pkg_data in data['packages']:
            pkg_elem = ET.SubElement(packages_elem, 'package')
            pkg_elem.set('name', pkg_data['name'])
            
            for key in ['version', 'license', 'homepage', 'supplier', 'root_path',
                       'total_files', 'used_files_count', 'usage_percentage', 'checksum']:
                elem = ET.SubElement(pkg_elem, key)
                elem.text = str(pkg_data.get(key, ''))
            
            used_files_elem = ET.SubElement(pkg_elem, 'used_files')
            for filepath in pkg_data.get('used_files', []):
                file_elem = ET.SubElement(used_files_elem, 'file')
                file_elem.text = filepath
        
        # Pretty print
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
        
        with open(output_path, 'w') as f:
            f.write(xml_str)
    
    # ========================================================================
    # XLSX REPORT
    # ========================================================================
    
    def _generate_xlsx(self, output_path: str) -> None:
        """Generate Excel report"""
        if not HAS_XLSX:
            raise ImportError("openpyxl required for XLSX: pip install openpyxl")
        
        wb = Workbook()
        
        # Summary sheet
        self._create_summary_sheet(wb)
        
        # Packages sheet
        self._create_packages_sheet(wb)
        
        # Files sheet
        self._create_files_sheet(wb)
        
        # Statistics sheet
        self._create_statistics_sheet(wb)
        
        wb.save(output_path)
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary worksheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Title
        ws['A1'] = 'BuildFileTracker Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = 'Generated:'
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A3'] = 'Backend:'
        ws['B3'] = self.tracker.active_backend.value if self.tracker.active_backend else 'N/A'
        
        # Statistics
        ws['A5'] = 'Statistics'
        ws['A5'].font = Font(size=14, bold=True)
        
        stats = self.tracker.stats
        row = 6
        for key, value in [
            ('Total Packages', stats.total_packages),
            ('Packages Used', stats.packages_used),
            ('Unique Files', stats.unique_files),
            ('Total Events', stats.total_events),
            ('Runtime (seconds)', f"{stats.runtime_seconds:.2f}"),
        ]:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
    
    def _create_packages_sheet(self, wb: Workbook) -> None:
        """Create packages worksheet"""
        ws = wb.create_sheet("Packages")
        
        # Header
        headers = ['Package', 'Version', 'License', 'Homepage', 'Total Files',
                  'Used Files', 'Usage %', 'Checksum']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for pkg in self.tracker.get_all_packages():
            ws.append([
                pkg.name,
                pkg.version,
                pkg.license,
                pkg.homepage,
                pkg.total_files,
                pkg.used_files_count,
                f"{pkg.usage_percentage:.2f}%",
                pkg.checksum
            ])
        
        # Adjust column widths
        for col, width in zip('ABCDEFGH', [30, 15, 20, 40, 12, 12, 12, 20]):
            ws.column_dimensions[col].width = width
    
    def _create_files_sheet(self, wb: Workbook) -> None:
        """Create files worksheet"""
        ws = wb.create_sheet("Used Files")
        
        ws.append(['Package', 'File Path', 'Access Types'])
        
        for pkg in self.tracker.get_all_packages():
            for filepath in sorted(pkg.used_files):
                access_types = ', '.join(sorted(self.tracker.file_access_map.get(filepath, [])))
                ws.append([pkg.name, filepath, access_types])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 20
    
    def _create_statistics_sheet(self, wb: Workbook) -> None:
        """Create statistics worksheet"""
        ws = wb.create_sheet("Statistics")
        
        ws['A1'] = 'Build Statistics'
        ws['A1'].font = Font(size=14, bold=True)
        
        stats = self.tracker.stats.to_dict()
        row = 3
        for key, value in stats.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
    
    # ========================================================================
    # SPDX JSON REPORT
    # ========================================================================
    
    def _generate_spdx_json(self, output_path: str) -> None:
        """Generate SPDX 2.3 JSON format"""
        doc_uuid = str(uuid.uuid4())
        
        spdx = {
            "spdxVersion": "SPDX-2.3",
            "dataLicense": "CC0-1.0",
            "SPDXID": "SPDXRef-DOCUMENT",
            "name": "BuildFileTracker-SBOM",
            "documentNamespace": f"https://buildfiletracker.local/{doc_uuid}",
            "creationInfo": {
                "created": datetime.now().isoformat() + "Z",
                "creators": ["Tool: BuildFileTracker-2.0.0"],
                "licenseListVersion": "3.21"
            },
            "packages": [],
            "files": [],
            "relationships": []
        }
        
        # Add packages
        for idx, pkg in enumerate(self.tracker.get_all_packages()):
            pkg_id = f"SPDXRef-Package-{idx}"
            
            spdx_pkg = {
                "SPDXID": pkg_id,
                "name": pkg.name,
                "versionInfo": pkg.version,
                "downloadLocation": pkg.homepage or "NOASSERTION",
                "supplier": f"Organization: {pkg.supplier}" if pkg.supplier else "NOASSERTION",
                "filesAnalyzed": True,
                "licenseConcluded": pkg.license if pkg.license != 'unknown' else "NOASSERTION",
                "licenseDeclared": pkg.license if pkg.license != 'unknown' else "NOASSERTION",
                "copyrightText": "NOASSERTION",
                "comment": f"Build usage: {pkg.usage_percentage:.2f}% ({pkg.used_files_count}/{pkg.total_files} files)",
                "checksums": [{
                    "algorithm": "SHA256",
                    "checksumValue": pkg.checksum if pkg.checksum else "0" * 64
                }]
            }
            
            spdx["packages"].append(spdx_pkg)
            
            # Add relationship
            spdx["relationships"].append({
                "spdxElementId": "SPDXRef-DOCUMENT",
                "relationshipType": "DESCRIBES",
                "relatedSpdxElement": pkg_id
            })
            
            # Add used files
            for file_idx, filepath in enumerate(sorted(pkg.used_files)):
                file_id = f"SPDXRef-File-{idx}-{file_idx}"
                
                spdx["files"].append({
                    "SPDXID": file_id,
                    "fileName": filepath,
                    "licenseConcluded": "NOASSERTION",
                    "copyrightText": "NOASSERTION"
                })
                
                spdx["relationships"].append({
                    "spdxElementId": pkg_id,
                    "relationshipType": "CONTAINS",
                    "relatedSpdxElement": file_id
                })
        
        with open(output_path, 'w') as f:
            json.dump(spdx, f, indent=2)
    
    # ========================================================================
    # SPDX TAG-VALUE REPORT
    # ========================================================================
    
    def _generate_spdx_tagvalue(self, output_path: str) -> None:
        """Generate SPDX Tag-Value format"""
        doc_uuid = str(uuid.uuid4())
        
        with open(output_path, 'w') as f:
            # Document info
            f.write("SPDXVersion: SPDX-2.3\n")
            f.write("DataLicense: CC0-1.0\n")
            f.write("SPDXID: SPDXRef-DOCUMENT\n")
            f.write("DocumentName: BuildFileTracker-SBOM\n")
            f.write(f"DocumentNamespace: https://buildfiletracker.local/{doc_uuid}\n")
            f.write("\n")
            
            # Creation info
            f.write(f"Creator: Tool: BuildFileTracker-2.0.0\n")
            f.write(f"Created: {datetime.now().isoformat()}Z\n")
            f.write("\n")
            
            # Packages
            for idx, pkg in enumerate(self.tracker.get_all_packages()):
                pkg_id = f"SPDXRef-Package-{idx}"
                
                f.write(f"PackageName: {pkg.name}\n")
                f.write(f"SPDXID: {pkg_id}\n")
                f.write(f"PackageVersion: {pkg.version}\n")
                f.write(f"PackageDownloadLocation: {pkg.homepage or 'NOASSERTION'}\n")
                f.write(f"FilesAnalyzed: true\n")
                f.write(f"PackageLicenseConcluded: {pkg.license if pkg.license != 'unknown' else 'NOASSERTION'}\n")
                f.write(f"PackageLicenseDeclared: {pkg.license if pkg.license != 'unknown' else 'NOASSERTION'}\n")
                f.write(f"PackageCopyrightText: NOASSERTION\n")
                f.write(f"PackageComment: Build usage: {pkg.usage_percentage:.2f}% ({pkg.used_files_count}/{pkg.total_files} files)\n")
                f.write("\n")
                
                # Relationship
                f.write(f"Relationship: SPDXRef-DOCUMENT DESCRIBES {pkg_id}\n")
                f.write("\n")
    
    # ========================================================================
    # CYCLONEDX JSON REPORT
    # ========================================================================
    
    def _generate_cyclonedx_json(self, output_path: str) -> None:
        """Generate CycloneDX 1.4 JSON format"""
        bom = {
            "bomFormat": "CycloneDX",
            "specVersion": "1.4",
            "serialNumber": f"urn:uuid:{uuid.uuid4()}",
            "version": 1,
            "metadata": {
                "timestamp": datetime.now().isoformat() + "Z",
                "tools": [{
                    "vendor": "BuildFileTracker",
                    "name": "BuildFileTracker",
                    "version": "2.0.0"
                }],
                "properties": [
                    {
                        "name": "buildfiletracker:backend",
                        "value": self.tracker.active_backend.value if self.tracker.active_backend else "none"
                    },
                    {
                        "name": "buildfiletracker:total_events",
                        "value": str(self.tracker.stats.total_events)
                    }
                ]
            },
            "components": []
        }
        
        # Add components (packages)
        for pkg in self.tracker.get_all_packages():
            component = {
                "type": "library",
                "name": pkg.name,
                "version": pkg.version,
                "hashes": [{
                    "alg": "SHA-256",
                    "content": pkg.checksum if pkg.checksum else "0" * 64
                }],
                "licenses": [{
                    "license": {
                        "id": pkg.license if pkg.license != 'unknown' else None,
                        "name": pkg.license if pkg.license != 'unknown' else "Unknown"
                    }
                }] if pkg.license else [],
                "properties": [
                    {
                        "name": "buildfiletracker:total_files",
                        "value": str(pkg.total_files)
                    },
                    {
                        "name": "buildfiletracker:used_files",
                        "value": str(pkg.used_files_count)
                    },
                    {
                        "name": "buildfiletracker:usage_percentage",
                        "value": f"{pkg.usage_percentage:.2f}"
                    },
                    {
                        "name": "buildfiletracker:root_path",
                        "value": pkg.root_path
                    }
                ]
            }
            
            if pkg.homepage:
                component["externalReferences"] = [{
                    "type": "website",
                    "url": pkg.homepage
                }]
            
            bom["components"].append(component)
        
        with open(output_path, 'w') as f:
            json.dump(bom, f, indent=2)
    
    # ========================================================================
    # CYCLONEDX XML REPORT
    # ========================================================================
    
    def _generate_cyclonedx_xml(self, output_path: str) -> None:
        """Generate CycloneDX XML format"""
        root = ET.Element('bom', xmlns="http://cyclonedx.org/schema/bom/1.4")
        root.set('serialNumber', f"urn:uuid:{uuid.uuid4()}")
        root.set('version', '1')
        
        # Metadata
        metadata = ET.SubElement(root, 'metadata')
        timestamp = ET.SubElement(metadata, 'timestamp')
        timestamp.text = datetime.now().isoformat() + "Z"
        
        # Components
        components = ET.SubElement(root, 'components')
        
        for pkg in self.tracker.get_all_packages():
            component = ET.SubElement(components, 'component', type='library')
            
            ET.SubElement(component, 'name').text = pkg.name
            ET.SubElement(component, 'version').text = pkg.version
            
            # Properties
            properties = ET.SubElement(component, 'properties')
            for key, value in [
                ('total_files', pkg.total_files),
                ('used_files', pkg.used_files_count),
                ('usage_percentage', f"{pkg.usage_percentage:.2f}")
            ]:
                prop = ET.SubElement(properties, 'property', name=f"buildfiletracker:{key}")
                prop.text = str(value)
        
        tree = ET.ElementTree(root)
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
    
    # ========================================================================
    # MARKDOWN REPORT
    # ========================================================================
    
    def _generate_markdown(self, output_path: str) -> None:
        """Generate Markdown report"""
        with open(output_path, 'w') as f:
            f.write("# BuildFileTracker Report\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
            f.write(f"**Backend:** {self.tracker.active_backend.value if self.tracker.active_backend else 'N/A'}  \n")
            f.write(f"**Runtime:** {self.tracker.stats.runtime_seconds:.2f} seconds\n\n")
            
            # Statistics
            f.write("## Statistics\n\n")
            f.write("| Metric | Value |\n")
            f.write("|--------|-------|\n")
            for key, value in self.tracker.stats.to_dict().items():
                f.write(f"| {key.replace('_', ' ').title()} | {value} |\n")
            f.write("\n")
            
            # Packages
            f.write("## Package Usage\n\n")
            for pkg in sorted(self.tracker.get_all_packages(), 
                            key=lambda p: p.usage_percentage, reverse=True):
                f.write(f"### {pkg.name} ({pkg.version})\n\n")
                f.write(f"- **License:** {pkg.license}\n")
                f.write(f"- **Total Files:** {pkg.total_files}\n")
                f.write(f"- **Used Files:** {pkg.used_files_count}\n")
                f.write(f"- **Usage:** {pkg.usage_percentage:.2f}%\n")
                
                if pkg.used_files:
                    f.write(f"\n**Files Used:**\n\n")
                    for filepath in sorted(pkg.used_files)[:20]:  # Limit to 20
                        f.write(f"- `{filepath}`\n")
                    if len(pkg.used_files) > 20:
                        f.write(f"- ... and {len(pkg.used_files) - 20} more\n")
                
                f.write("\n")
    
    # ========================================================================
    # HTML REPORT
    # ========================================================================
    
    def _generate_html(self, output_path: str) -> None:
        """Generate interactive HTML report"""
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>BuildFileTracker Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
               margin: 0; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; 
                      border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 3px solid #0066cc; padding-bottom: 10px; }}
        h2 {{ color: #0066cc; margin-top: 30px; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                 gap: 20px; margin: 20px 0; }}
        .stat-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }}
        .stat-value {{ font-size: 2em; font-weight: bold; color: #0066cc; }}
        .stat-label {{ color: #666; margin-top: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #0066cc; color: white; }}
        tr:hover {{ background: #f5f5f5; }}
        .progress-bar {{ background: #e0e0e0; border-radius: 4px; height: 20px; width: 200px; }}
        .progress-fill {{ background: #4caf50; height: 100%; border-radius: 4px; }}
        .badge {{ display: inline-block; padding: 4px 8px; border-radius: 4px; 
                 font-size: 0.85em; background: #e3f2fd; color: #1976d2; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 BuildFileTracker Report</h1>
        <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>Backend:</strong> {self.tracker.active_backend.value if self.tracker.active_backend else 'N/A'}</p>
        
        <h2>📊 Statistics</h2>
        <div class="stats">
            <div class="stat-card">
                <div class="stat-value">{self.tracker.stats.total_packages}</div>
                <div class="stat-label">Total Packages</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{self.tracker.stats.packages_used}</div>
                <div class="stat-label">Packages Used</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{self.tracker.stats.unique_files}</div>
                <div class="stat-label">Files Accessed</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{self.tracker.stats.runtime_seconds:.2f}s</div>
                <div class="stat-label">Runtime</div>
            </div>
        </div>
        
        <h2>📦 Package Usage</h2>
        <table>
            <thead>
                <tr>
                    <th>Package</th>
                    <th>Version</th>
                    <th>License</th>
                    <th>Usage</th>
                    <th>Files</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for pkg in sorted(self.tracker.get_all_packages(), 
                         key=lambda p: p.usage_percentage, reverse=True):
            html += f"""
                <tr>
                    <td><strong>{pkg.name}</strong></td>
                    <td>{pkg.version}</td>
                    <td><span class="badge">{pkg.license}</span></td>
                    <td>
                        <div class="progress-bar">
                            <div class="progress-fill" style="width: {min(pkg.usage_percentage, 100)}%"></div>
                        </div>
                        {pkg.usage_percentage:.2f}%
                    </td>
                    <td>{pkg.used_files_count} / {pkg.total_files}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
        
        <footer style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; 
                       text-align: center; color: #666;">
            <p>Generated by BuildFileTracker v2.0.0</p>
        </footer>
    </div>
</body>
</html>
"""
        
        with open(output_path, 'w') as f:
            f.write(html)
